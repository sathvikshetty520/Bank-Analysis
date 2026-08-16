"""
Report Generation
--------------------
Generates the two deliverables from Requirement 6: a full-detail Excel
workbook (raw ledger + every analysis sheet, for further work in a
spreadsheet) and a summary PDF (the human-readable investigation
report an officer would actually read/print/file).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="14181A", end_color="14181A", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
FLAG_FONT = Font(name="Arial", size=10, color="A6392E", bold=True)


def _style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws, num_cols, width=18):
    for col in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def generate_excel_report(transactions, round_trips, accumulation_accounts, trails, output_path):
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Bank Statement Investigation Report"
    ws["A1"].font = Font(name="Arial", size=16, bold=True)
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    _style_header_row(ws, 3, 2)

    accounts = sorted(set(t.account_id for t in transactions))
    dup_count = sum(1 for t in transactions if t.is_duplicate)
    mismatch_count = sum(1 for t in transactions if t.balance_mismatch)
    confirmed_rev = sum(1 for t in transactions if t.reversal_confidence == "confirmed")
    possible_rev = sum(1 for t in transactions if t.reversal_confidence == "possible")

    summary_rows = [
        ("Accounts in investigation", ", ".join(accounts)),
        ("Total transactions", len(transactions)),
        ("Duplicates flagged", dup_count),
        ("Balance-chain mismatches", mismatch_count),
        ("Reversed - confirmed (bank-stated)", confirmed_rev),
        ("Reversed - possible (amount match, needs review)", possible_rev),
        ("Round-trip patterns detected", len(round_trips)),
        ("Accumulation leads identified", len(accumulation_accounts)),
    ]
    for i, (label, val) in enumerate(summary_rows, start=4):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        ws.cell(row=i, column=2, value=val).font = BODY_FONT
    _autosize_columns(ws, 2, width=45)

    # --- Sheet 2: Transaction Ledger ---
    ws2 = wb.create_sheet("Transaction Ledger")
    headers = ["Account", "Date", "Narration", "Ref No", "Debit", "Credit", "Balance",
               "Duplicate", "Reversed", "Reversal Confidence", "Balance Mismatch"]
    for c, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, 1, len(headers))

    row = 2
    for t in sorted(transactions, key=lambda t: (t.account_id, t.date, t.source_row)):
        ws2.cell(row=row, column=1, value=t.account_id).font = BODY_FONT
        ws2.cell(row=row, column=2, value=str(t.date)).font = BODY_FONT
        ws2.cell(row=row, column=3, value=t.narration).font = BODY_FONT
        ws2.cell(row=row, column=4, value=t.ref_no or "").font = BODY_FONT
        ws2.cell(row=row, column=5, value=t.debit if t.debit else None).font = BODY_FONT
        ws2.cell(row=row, column=6, value=t.credit if t.credit else None).font = BODY_FONT
        ws2.cell(row=row, column=7, value=t.balance).font = BODY_FONT
        ws2.cell(row=row, column=8, value="YES" if t.is_duplicate else "").font = FLAG_FONT
        ws2.cell(row=row, column=9, value="YES" if t.is_reversed_transaction else "").font = FLAG_FONT
        ws2.cell(row=row, column=10, value=t.reversal_confidence or "").font = BODY_FONT
        ws2.cell(row=row, column=11, value="YES" if t.balance_mismatch else "").font = FLAG_FONT
        row += 1

    # Real SUM formulas, not hardcoded Python totals - sheet recalculates if edited
    last_row = row - 1
    ws2.cell(row=row, column=4, value="TOTAL").font = Font(name="Arial", bold=True)
    ws2.cell(row=row, column=5, value=f"=SUM(E2:E{last_row})").font = Font(name="Arial", bold=True)
    ws2.cell(row=row, column=6, value=f"=SUM(F2:F{last_row})").font = Font(name="Arial", bold=True)
    _autosize_columns(ws2, len(headers), width=16)
    ws2.column_dimensions["C"].width = 50

    # --- Sheet 3: Round Trips ---
    ws3 = wb.create_sheet("Round Trips")
    rt_headers = ["Round Trip #", "From Account", "To Account", "Date", "Amount", "Confirmed Link"]
    for c, h in enumerate(rt_headers, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header_row(ws3, 1, len(rt_headers))
    row = 2
    for i, rt in enumerate(round_trips, start=1):
        for e in rt["edges"]:
            ws3.cell(row=row, column=1, value=i).font = BODY_FONT
            ws3.cell(row=row, column=2, value=e["from"]).font = BODY_FONT
            ws3.cell(row=row, column=3, value=e["to"]).font = BODY_FONT
            ws3.cell(row=row, column=4, value=str(e["date"])).font = BODY_FONT
            ws3.cell(row=row, column=5, value=e["amount"]).font = BODY_FONT
            ws3.cell(row=row, column=6, value="YES" if e.get("confirmed_link") else "").font = BODY_FONT
            row += 1
    _autosize_columns(ws3, len(rt_headers), width=20)
    if not round_trips:
        ws3.cell(row=2, column=1, value="No round-trip patterns detected in this investigation.").font = BODY_FONT

    # --- Sheet 4: Accumulation Leads ---
    ws4 = wb.create_sheet("Accumulation Leads")
    ws4.cell(row=1, column=1, value="Account / Counterparty")
    ws4.cell(row=1, column=2, value="Net Accumulated (Received - Sent)")
    _style_header_row(ws4, 1, 2)
    for i, a in enumerate(accumulation_accounts, start=2):
        ws4.cell(row=i, column=1, value=a["account"]).font = BODY_FONT
        ws4.cell(row=i, column=2, value=a["net_accumulated"]).font = BODY_FONT
    _autosize_columns(ws4, 2, width=30)

    # --- Sheet 5: Money Trail ---
    ws5 = wb.create_sheet("Money Trail")
    mt_headers = ["Source Credit Date", "Source Credit Narration", "Source Amount",
                  "Fully Traced", "Step Date", "Step Narration", "Amount From This Credit"]
    for c, h in enumerate(mt_headers, start=1):
        ws5.cell(row=1, column=c, value=h)
    _style_header_row(ws5, 1, len(mt_headers))
    row = 2
    for tr in trails:
        sc = tr["source_credit"]
        if not tr["trail"]:
            ws5.cell(row=row, column=1, value=str(sc["date"])).font = BODY_FONT
            ws5.cell(row=row, column=2, value=sc["narration"]).font = BODY_FONT
            ws5.cell(row=row, column=3, value=sc["amount"]).font = BODY_FONT
            ws5.cell(row=row, column=4, value="YES" if tr["fully_traced"] else "NO").font = BODY_FONT
            row += 1
        for step in tr["trail"]:
            ws5.cell(row=row, column=1, value=str(sc["date"])).font = BODY_FONT
            ws5.cell(row=row, column=2, value=sc["narration"]).font = BODY_FONT
            ws5.cell(row=row, column=3, value=sc["amount"]).font = BODY_FONT
            ws5.cell(row=row, column=4, value="YES" if tr["fully_traced"] else "NO").font = BODY_FONT
            ws5.cell(row=row, column=5, value=str(step["date"])).font = BODY_FONT
            ws5.cell(row=row, column=6, value=step["narration"]).font = BODY_FONT
            ws5.cell(row=row, column=7, value=step["amount_from_this_credit"]).font = BODY_FONT
            row += 1
    _autosize_columns(ws5, len(mt_headers), width=22)
    ws5.column_dimensions["B"].width = 40
    ws5.column_dimensions["F"].width = 40

    wb.save(output_path)
    return output_path


def generate_pdf_report(transactions, round_trips, accumulation_accounts, trails, output_path,
                         max_flagged_rows=60, max_trail_rows=15):
    """
    Human-readable summary report - not the full ledger (that's what the
    Excel export is for). Investigators read this to decide where to
    look next, not to audit every row.
    """
    doc = SimpleDocTemplate(output_path, pagesize=letter,
                             topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=20)
    heading_style = styles["Heading2"]
    body_style = styles["Normal"]
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    # Small style for table cell content - wrapping plain strings in Paragraph
    # objects like this is required for ReportLab to wrap long text within a
    # column width; plain strings in Table cells do NOT wrap and instead
    # overflow/overlap into neighboring cells (the bug this fixes).
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)
    cell_style_white = ParagraphStyle("CellWhite", parent=cell_style, textColor=colors.white, fontName="Helvetica-Bold")

    def cell(text):
        return Paragraph(str(text), cell_style)

    def header_cell(text):
        return Paragraph(str(text), cell_style_white)

    story = []

    # --- Cover / Summary ---
    story.append(Paragraph("Bank Statement Investigation Report", title_style))
    story.append(Spacer(1, 12))

    accounts = sorted(set(t.account_id for t in transactions))
    dup_count = sum(1 for t in transactions if t.is_duplicate)
    mismatch_count = sum(1 for t in transactions if t.balance_mismatch)
    confirmed_rev = sum(1 for t in transactions if t.reversal_confidence == "confirmed")
    possible_rev = sum(1 for t in transactions if t.reversal_confidence == "possible")

    summary_data = [
        [cell("Accounts in investigation"), cell(", ".join(accounts))],
        [cell("Total transactions analyzed"), cell(str(len(transactions)))],
        [cell("Duplicates flagged"), cell(str(dup_count))],
        [cell("Balance-chain mismatches"), cell(str(mismatch_count))],
        [cell("Reversed - confirmed (bank-stated)"), cell(str(confirmed_rev))],
        [cell("Reversed - possible (amount match)"), cell(str(possible_rev))],
        [cell("Round-trip patterns detected"), cell(str(len(round_trips)))],
        [cell("Accumulation leads identified"), cell(str(len(accumulation_accounts)))],
    ]
    t = Table(summary_data, colWidths=[3 * inch, 3.5 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F0EEE6")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8D3C4")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "This report summarizes automated analysis of the uploaded bank statement(s). "
        "\u201cPossible\u201d flags are heuristic signals (matching amounts and, on larger statements, "
        "matching counterparties) and are not confirmed findings - they require manual review. "
        "The full transaction ledger is available in the accompanying Excel export.",
        note_style
    ))
    story.append(PageBreak())

    # --- Round Trips ---
    story.append(Paragraph("Round-Trip Patterns", heading_style))
    if round_trips:
        for i, rt in enumerate(round_trips, start=1):
            chain = " \u2192 ".join(rt["accounts_involved"] + [rt["accounts_involved"][0]])
            story.append(Paragraph(f"<b>Round Trip {i}:</b> {chain} &nbsp; ({rt['span_days']} day span)", body_style))
            rt_data = [[header_cell("From"), header_cell("To"), header_cell("Date"),
                        header_cell("Amount"), header_cell("Confirmed Link")]]
            for e in rt["edges"]:
                rt_data.append([cell(e["from"]), cell(e["to"]), cell(str(e["date"])),
                                 cell(f"{e['amount']:,.2f}"),
                                 cell("Yes" if e.get("confirmed_link") else "No")])
            rt_table = Table(rt_data, colWidths=[1.3*inch, 1.3*inch, 1*inch, 1.2*inch, 1.2*inch])
            rt_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#A6392E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8D3C4")),
            ]))
            story.append(rt_table)
            story.append(Spacer(1, 10))
    else:
        story.append(Paragraph(
            "No round-trip patterns detected. Note: this requires statements from multiple "
            "linked accounts to be uploaded into the same investigation session - a single "
            "account's statement can only show money leaving or arriving, not whether it "
            "later returns through another account.", body_style))
    story.append(Spacer(1, 16))

    # --- Accumulation Leads ---
    story.append(Paragraph("Top Accumulation Leads", heading_style))
    if accumulation_accounts:
        lead_data = [[header_cell("Account / Counterparty"), header_cell("Net Accumulated")]]
        for a in accumulation_accounts:
            lead_data.append([cell(a["account"]), cell(f"Rs. {a['net_accumulated']:,.2f}")])
        lead_table = Table(lead_data, colWidths=[3.5 * inch, 2.5 * inch])
        lead_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14181A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8D3C4")),
        ]))
        story.append(lead_table)
    else:
        story.append(Paragraph("No accumulation leads identified.", body_style))
    story.append(PageBreak())

    # --- Flagged Transactions (capped, not the full ledger) ---
    story.append(Paragraph("Flagged Transactions", heading_style))
    flagged = [t for t in transactions if t.is_duplicate or t.is_reversed_transaction or t.balance_mismatch]
    story.append(Paragraph(
        f"{len(flagged)} transaction(s) flagged out of {len(transactions)} total. "
        f"Showing up to {max_flagged_rows} below; see Excel export for the complete list.",
        note_style
    ))
    story.append(Spacer(1, 6))
    if flagged:
        flag_data = [[header_cell("Date"), header_cell("Narration"), header_cell("Amount"), header_cell("Flags")]]
        for t in flagged[:max_flagged_rows]:
            flags = []
            if t.is_duplicate:
                flags.append("Duplicate")
            if t.is_reversed_transaction:
                flags.append(f"Reversed ({t.reversal_confidence})")
            if t.balance_mismatch:
                flags.append("Balance break")
            amt = t.debit if t.debit else t.credit
            flag_data.append([cell(str(t.date)), cell((t.narration or "")[:80]),
                               cell(f"{amt:,.2f}"), cell(", ".join(flags))])
        flag_table = Table(flag_data, colWidths=[0.8*inch, 3*inch, 1*inch, 1.7*inch])
        flag_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14181A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8D3C4")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(flag_table)
    else:
        story.append(Paragraph("No flagged transactions.", body_style))

    doc.build(story)
    return output_path